#read several txt file and fill each line to spreadsheet cells
#every column represent a file
#Usage: python txtToSpreadsheet.py FILE1  FILE2 FILE3 ...

import openpyxl, sys

def txt2Spreadsheet(filenameList):
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    titleFont = openpyxl.styles.Font(bold=True, italic=True)
    
    for fIdx, filename in enumerate(filenameList):
        #fill in filename to raw 1
        sheet.cell(row=1, column=fIdx+1).value = filename.split('.txt')[0]
        sheet.cell(row=1, column=fIdx+1).font= titleFont
        
        with open(filename, 'r') as file:
            for rIdx, line in enumerate(file.readlines()):
                #rIdx start from 0, and first row already fill in filename
                sheet.cell(row=rIdx+2, column=fIdx+1).value = line
    
    wb.save('txt2Spreadsheet.xlsx')

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python txtToSpreadsheet.py FILE1  FILE2 FILE3 ...')
    else:
        #check if all files are .txt
        filenameList = sys.argv[1:]
        for filename in filenameList:
            if not filename.endswith('.txt'):
                print('Only support .txt file.')
                exit()
        
        txt2Spreadsheet(filenameList)
