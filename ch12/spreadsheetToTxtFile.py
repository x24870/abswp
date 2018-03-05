#Convet spreadsheet to txt files
#Each column is a file, first row is file name, left rows are content.
#Usage: python spreadsheetToTxtFile.py FILE_NAME

import openpyxl, sys

def spreadsheet2Txt(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    for col in range(1, sheet.max_column+1):
        #check if this column is empty
        if sheet.cell(row=1, column=col).value != None:
            txtFilename = str(sheet.cell(row=1, column=col).value) + '.txt'
            print('Create file:', txtFilename)
            with open(txtFilename, 'w') as file:
                for row in range(2, sheet.max_row+1):
                    string = str(sheet.cell(row=row, column=col).value)
                    #check if this cell is empty, if empty the cell.value is 'None'
                    #issue: what if the real value is 'None'?
                    if string != 'None':
                        if not string.endswith('\n'): string += '\n'
                        print(type(string), string)
                        file.writelines(string)
        
    
if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python spreadsheetToTxtFile.py FILE_NAME')
    elif not sys.argv[1].endswith('.xlsx'):
        print('Only support .xlsx file.')
    else:
        spreadsheet2Txt(sys.argv[1])