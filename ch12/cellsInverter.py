#invert columns and rows of ccells in sheet
#Usage: python cellsInverter.py FILENAME

import openpyxl, sys

def cells_invert(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    #write to new sheet
    newSheet = wb.create_sheet(title='inverted')

    for row in range(1, sheet.max_row+1):
        for col in range(1, sheet.max_column+1):
            newSheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value
            
    wb.save('inverted_' + filename)
    
if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python cellsInverter.py FILENAME')
    elif not sys.argv[1].endswith('.xlsx'):
        print('Only support .xlsx file')
    else:
        cells_invert(sys.argv[1])